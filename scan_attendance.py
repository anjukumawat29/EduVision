import cv2
import os
import sys
import time
import pickle
import openpyxl
from datetime import datetime
from zoneinfo import ZoneInfo

# ── macOS GUI fix ─────────────────────────────────────────────────
try:
    from macos_display_helper import enable_macos_display, safe_imshow, safe_waitkey, safe_destroyall
    enable_macos_display()
except:
    # Fallback if helper not available
    def safe_imshow(w, i):
        try:
            cv2.imshow(w, i)
        except:
            pass
    def safe_waitkey(d=1):
        try:
            return (cv2.waitKey(d) & 0xFF) == ord('q')
        except:
            return False
    def safe_destroyall():
        try:
            cv2.destroyAllWindows()
        except:
            pass

# ── Args ─────────────────────────────
model_file  = sys.argv[1]
labels_file = sys.argv[2]
excel_file  = sys.argv[3]
duration    = int(sys.argv[4]) if len(sys.argv) > 4 else 20

CONFIDENCE_THRESHOLD = 85

# ── Load recognizer ──────────────────
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(model_file)

with open(labels_file, "rb") as f:
    label_map = pickle.load(f)

# ── Haar Cascade (face detection) ────
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# ── Camera ───────────────────────────
cap = cv2.VideoCapture(0)

if not cap.isOpened():
    print("[scan] ❌ Camera failed to open")
    exit()

cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # Prevent lag

cv2.namedWindow("Attendance Scan", cv2.WINDOW_NORMAL)
cv2.resizeWindow("Attendance Scan", 1200, 800)

print("[scan] Camera started")

# Quick warmup — 5 frames is enough for macOS AVFoundation
for _ in range(5):
    cap.read()

# ── Scan loop ────────────────────────
marked = {}
start = time.time()

while (time.time() - start) < duration:

    ret, frame = cap.read()
    if not ret or frame is None:
        continue

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)

    # MULTI-FACE DETECTION
    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(60, 60)
    )

    for (x, y, w, h) in faces:

        face_crop = gray[y:y+h, x:x+w]

        try:
            face_crop = cv2.resize(face_crop, (200, 200))
        except:
            continue

        # ── Recognition ─────────────────
        label, conf = recognizer.predict(face_crop)

        if conf < CONFIDENCE_THRESHOLD and label in label_map:
            name = label_map[label]
            color = (0, 255, 0)

            if name not in marked:
                marked[name] = conf
                print("[scan] Present:", name)

        else:
            name = "Unknown"
            color = (0, 0, 255)

        # ── Draw box ───────────────────
        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, name, (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    # ── HUD ───────────────────────────
    remaining = int(duration - (time.time() - start))
    cv2.putText(frame, f"{remaining}s left",
                (20, 40),
                cv2.FONT_HERSHEY_SIMPLEX, 1,
                (255, 255, 255), 2)

    if marked:
        cv2.putText(frame, "Marked: " + ", ".join(marked.keys()),
                    (20, frame.shape[0] - 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                    (0, 255, 0), 2)

    # ── Display ───────────────────────
    safe_imshow("Attendance Scan", frame)

    if safe_waitkey(1):
        print("[scan] Stopped early")
        break

cap.release()
safe_destroyall()

# ── Save IST time ────────────────────
now = datetime.now(ZoneInfo("Asia/Kolkata"))

# ── Save Excel ───────────────────────
if marked:
    wb = openpyxl.load_workbook(excel_file) if os.path.exists(excel_file) else openpyxl.Workbook()
    ws = wb.active

    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(["Name", "Date", "Time", "Status"])

    for name in marked:
        ws.append([
            name,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            "Present"
        ])

    wb.save(excel_file)
    print("[scan] Attendance saved")

else:
    print("[scan] No students recognized")

# ── Output for Django ────────────────
print("RESULT:" + ",".join(marked.keys()))