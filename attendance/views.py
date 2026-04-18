import json
import cv2
import os
import sys
import pickle
import shutil
import subprocess
import openpyxl
import numpy as np
import time
from datetime import datetime
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from ultralytics import YOLO
from django.http import HttpResponse
from django.contrib.auth.models import User

# ── Config ────────────────────────────────────────────────────────
yolo = YOLO("yolov8n.pt")

DATASET_DIR  = os.path.join(settings.BASE_DIR, "dataset")
MODEL_FILE   = os.path.join(settings.BASE_DIR, "ml_models", "lbph_model.yml")
LABELS_FILE  = os.path.join(settings.BASE_DIR, "ml_models", "labels.pkl")
EXCEL_FILE   = os.path.join(settings.BASE_DIR, "attendance.xlsx")

CONFIDENCE_THRESHOLD = 85
MIN_PHOTOS           = 10

_CASCADE_PATH = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
face_cascade  = cv2.CascadeClassifier(_CASCADE_PATH)


# ── Camera (attendance/behavior only — NOT register) ──────────────
def open_camera():
    cam = cv2.VideoCapture(0, cv2.CAP_AVFOUNDATION)
    if not cam.isOpened():
        cam = cv2.VideoCapture(0)
    cam.set(cv2.CAP_PROP_FRAME_WIDTH,  640)
    cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    # Quick warmup — 5 frames is enough for macOS AVFoundation
    for _ in range(5):
        cam.read()
    return cam


def safe_imshow(title, frame):
    try:
        cv2.imshow(title, frame)
        return True
    except Exception as e:
        print("imshow error:", e)
        return False


# ── Face detection ────────────────────────────────────────────────
def detect_and_crop_face(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)

    faces = face_cascade.detectMultiScale(
        gray, scaleFactor=1.05, minNeighbors=5,
        minSize=(60, 60), flags=cv2.CASCADE_SCALE_IMAGE,
    )
    if len(faces) > 0:
        x, y, w, h = max(faces, key=lambda r: r[2] * r[3])
        return cv2.resize(gray[y:y+h, x:x+w], (200, 200)), (x, y, x+w, y+h)

    results = yolo(frame, verbose=False)[0]
    for box in results.boxes:
        if int(box.cls[0]) != 0:
            continue
        px1, py1, px2, py2 = map(int, box.xyxy[0])
        px1, py1 = max(0, px1), max(0, py1)
        px2, py2 = min(frame.shape[1], px2), min(frame.shape[0], py2)
        roi = gray[py1:py2, px1:px2]
        if roi.size == 0:
            continue
        inner = face_cascade.detectMultiScale(
            roi, scaleFactor=1.05, minNeighbors=4, minSize=(40, 40)
        )
        if len(inner) > 0:
            ix, iy, iw, ih = max(inner, key=lambda r: r[2] * r[3])
            ax, ay = px1 + ix, py1 + iy
            return cv2.resize(gray[ay:ay+ih, ax:ax+iw], (200, 200)), (ax, ay, ax+iw, ay+ih)

    return None, None


# ── Helpers ───────────────────────────────────────────────────────
def get_students():
    if not os.path.exists(DATASET_DIR):
        return []
    return sorted([
        s for s in os.listdir(DATASET_DIR)
        if os.path.isdir(os.path.join(DATASET_DIR, s))
    ])


def build_and_train():
    faces, labels, label_map = [], [], {}
    label_id = 0

    for student in get_students():
        folder        = os.path.join(DATASET_DIR, student)
        student_faces = []

        for img_file in os.listdir(folder):
            path = os.path.join(folder, img_file)
            img  = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
            if img is None or img.size == 0:
                continue
            if np.mean(img) < 20:
                continue
            student_faces.append(cv2.resize(img, (200, 200)))

        if len(student_faces) < MIN_PHOTOS:
            print(f"[train] Skipping '{student}' - only {len(student_faces)} photos (need {MIN_PHOTOS})")
            continue

        for face in student_faces:
            faces.append(face)
            labels.append(label_id)

        label_map[label_id] = student
        label_id += 1

    if not faces:
        return False, f"No valid images. Each student needs at least {MIN_PHOTOS} photos."

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(faces, np.array(labels))

    os.makedirs(os.path.dirname(MODEL_FILE), exist_ok=True)
    recognizer.save(MODEL_FILE)

    with open(LABELS_FILE, "wb") as f:
        pickle.dump(label_map, f)

    return True, f"Model trained on {label_id} student(s) with {len(faces)} total images."


def load_recognizer():
    if not os.path.exists(MODEL_FILE) or not os.path.exists(LABELS_FILE):
        return None, None
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read(MODEL_FILE)
    with open(LABELS_FILE, "rb") as f:
        label_map = pickle.load(f)
    return recognizer, label_map


def append_excel(rows):
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time", "Status"])
    for row in rows:
        ws.append(row)
    wb.save(EXCEL_FILE)


def read_recent_excel(n=10):
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data.append({
            "name":   row[0] or "-",
            "date":   row[1] or "-",
            "time":   row[2] or "-",
            "status": row[3] or "-",
        })
    return data[-n:][::-1]


# ════════════════════════════════════════════════════════════════════
#  VIEWS
# ════════════════════════════════════════════════════════════════════

def home(request):
    """Teacher dashboard with profile card, today's attendance, and behavior insights."""

    # ── Behavior alerts ───────────────────────────────────────────────
    behavior_alerts = []
    if "behavior_log" in request.session:
        log = request.session.get("behavior_log", {})
        using_phone   = log.get("using_phone", 0)
        distracted    = log.get("distracted", 0)
        total         = log.get("total", 0)
        if total > 0:
            phone_pct      = (using_phone / total) * 100
            distracted_pct = (distracted  / total) * 100
            if phone_pct > 15:
                behavior_alerts.append({
                    "type":     "phone",
                    "message":  f"High phone usage detected ({phone_pct:.0f}%)",
                    "severity": "danger" if phone_pct > 30 else "warning",
                    "count":    using_phone,
                })
            if distracted_pct > 25:
                behavior_alerts.append({
                    "type":     "distracted",
                    "message":  f"Students distracted for {distracted_pct:.0f}% of session",
                    "severity": "warning",
                    "count":    distracted,
                })

    # ── Today's attendance summary ────────────────────────────────────
    today_present = []
    today_absent  = []
    today_str = datetime.now().strftime("%Y-%m-%d")
    all_students = get_students()

    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            present_today = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                name, date, _time, status = row[0], row[1], row[2], row[3]
                if str(date) == today_str and status == "Present":
                    present_today.add(str(name))
            today_present = sorted(present_today)
            today_absent  = [s for s in all_students if s not in present_today]
        except Exception:
            pass

    # ── Recent scans (last 5 attendance entries) ──────────────────────
    recent_scans = read_recent_excel(n=5)

    return render(request, "home.html", {
        "student_count":   len(all_students),
        "model_ready":     os.path.exists(MODEL_FILE),
        "min_photos":      MIN_PHOTOS,
        "behavior_alerts": behavior_alerts,
        "today_present":   today_present,
        "today_absent":    today_absent,
        "today_present_count": len(today_present),
        "today_absent_count":  len(today_absent),
        "recent_scans":    recent_scans,
    })


def student_list(request):
    students = []
    for s in get_students():
        folder = os.path.join(DATASET_DIR, s)
        count  = len([f for f in os.listdir(folder) if f.lower().endswith((".jpg", ".png", ".jpeg"))])
        students.append({"name": s, "photos": count})
    return render(request, "students.html", {"students": students})


def delete_student(request, name):
    folder = os.path.join(DATASET_DIR, name)
    if os.path.exists(folder):
        shutil.rmtree(folder)
        if get_students():
            build_and_train()
        else:
            for f in [MODEL_FILE, LABELS_FILE]:
                if os.path.exists(f):
                    os.remove(f)
        messages.success(request, f"'{name}' deleted. Retrain the model.")
    else:
        messages.error(request, "Student not found.")
    return redirect("students")


# ── Register via subprocess ───────────────────────────────────────

def register_student(request):
    if request.method != "POST":
        return render(request, "register.html")

    name = request.POST.get("name", "").strip()
    password = request.POST.get("password", "").strip()
    role = request.POST.get("role", "student")

    if not name:
        messages.error(request, "Name is required.")
        return render(request, "register.html")

    if not password:
        messages.error(request, "Password is required.")
        return render(request, "register.html")

    if User.objects.filter(username=name).exists():
        messages.warning(request, f"Username '{name}' already exists. Please log in.")
        return render(request, "register.html")

    user = User.objects.create_user(username=name, password=password)
    user.userprofile.is_teacher = (role == "teacher")
    user.userprofile.save()

    messages.success(request, f"Account created for '{name}'! Please sign in.")
    return redirect("login")


def capture_faces_view(request):
    """Student-only: launch face capture subprocess from the student dashboard."""
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method != "POST":
        return redirect("student_dashboard")

    name = request.user.username
    try:
        count = max(10, min(40, int(request.POST.get("count", "20"))))
    except ValueError:
        count = 20

    script_path = os.path.join(settings.BASE_DIR, "capture_faces.py")
    if not os.path.exists(script_path):
        messages.error(request, "capture_faces.py not found in project root.")
        return redirect("student_dashboard")

    try:
        subprocess.Popen([sys.executable, script_path, name, str(count)])
        messages.success(request,
            f"Camera launching — capturing {count} face photos for '{name}'. "
            "A window will open shortly. Keep still and face the camera!")
    except Exception as e:
        messages.error(request, f"Camera error: {e}")

    return redirect("student_dashboard")

# ── Train ─────────────────────────────────────────────────────────
def bulk_train(request):
    ok, msg = build_and_train()
    messages.success(request, msg) if ok else messages.error(request, msg)
    return redirect("home")


def retrain(request):
    return bulk_train(request)


# ── Attendance ────────────────────────────────────────────────────
ATTENDANCE_STEPS = [
    "Webcam opens for 20 seconds.",
    "Haar Cascade finds faces; LBPH matches against trained profiles.",
    "Each matched student is marked Present in attendance.xlsx.",
    "Press Q to end the scan early.",
]


def attendance_page(request):
    return render(request, "attendance.html", {
        "model_ready": os.path.exists(MODEL_FILE),
        "recent":      read_recent_excel(),
        "steps":       ATTENDANCE_STEPS,
    })


def mark_attendance(request):
    """
    Runs scan_attendance.py in a subprocess — same reason as register_student:
    cv2.imshow crashes with 'Unknown C++ exception' on macOS when called from
    a Django worker thread instead of the process main thread.
    We wait (blocking) for the subprocess to finish so we can read the result
    and show the correct flash message.
    """
    recognizer, label_map = load_recognizer()
    if recognizer is None:
        messages.error(request, "Model not trained yet. Register students and click Retrain Now.")
        return redirect("attendance_page")

    script_path = os.path.join(settings.BASE_DIR, "scan_attendance.py")
    if not os.path.exists(script_path):
        messages.error(request,
            "scan_attendance.py not found in project root. Place it next to manage.py.")
        return redirect("attendance_page")

    try:
        # macOS: pass environment to properly initialize GUI
        env = os.environ.copy()
        
        result = subprocess.run(
            [sys.executable, script_path,
             MODEL_FILE, LABELS_FILE, EXCEL_FILE, "20"],
            capture_output=True, text=True, timeout=60,
            env=env,
        )
        # Script prints "RESULT:name1,name2" on the last line
        marked = []
        for line in result.stdout.strip().splitlines():
            if line.startswith("RESULT:"):
                names = line[len("RESULT:"):].strip()
                marked = [n for n in names.split(",") if n]
                break

        if marked:
            messages.success(request, f"Marked Present: {', '.join(marked)}")
        else:
            messages.warning(request, "No student recognised. Try retraining or improving lighting.")

        if result.returncode != 0 and result.stderr:
            print("[attendance stderr]", result.stderr[:500])

    except subprocess.TimeoutExpired:
        messages.error(request, "Scan timed out. Please try again.")
    except Exception as e:
        messages.error(request, f"Scan failed: {e}")

    return redirect("attendance_page")

def export_attendance(request):
    if not os.path.exists(EXCEL_FILE):
        return HttpResponse("No attendance data found.", status=404)

    with open(EXCEL_FILE, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
        return response


def student_photo(request, username, photo_name):
    """Serve student's captured photo"""
    # Security: only allow authenticated users to view their own photos or teachers to view all
    if not request.user.is_authenticated:
        return HttpResponse("Unauthorized", status=401)
    
    is_own_photo = request.user.username.lower() == username.lower()
    is_teacher = hasattr(request.user, 'userprofile') and request.user.userprofile.is_teacher
    
    if not (is_own_photo or is_teacher):
        return HttpResponse("Forbidden", status=403)
    
    # Security: prevent directory traversal
    if ".." in photo_name or "/" in photo_name:
        return HttpResponse("Invalid filename", status=400)
    
    # Only allow image files
    if not photo_name.lower().endswith(('.jpg', '.jpeg', '.png')):
        return HttpResponse("Invalid file type", status=400)
    
    photo_path = os.path.join(DATASET_DIR, username, photo_name)
    
    if not os.path.exists(photo_path):
        return HttpResponse("Photo not found", status=404)
    
    try:
        with open(photo_path, 'rb') as f:
            if photo_name.lower().endswith('.png'):
                content_type = 'image/png'
            else:
                content_type = 'image/jpeg'
            
            response = HttpResponse(f.read(), content_type=content_type)
            response['Cache-Control'] = 'max-age=3600'
            return response
    except Exception as e:
        return HttpResponse(f"Error reading file: {e}", status=500)